import os
from openpyxl import load_workbook
from django.conf import settings
from tableapp.models import CustomUser, Table, Booking
from datetime import datetime, time
from django.core.management.base import BaseCommand

class Command(BaseCommand):
    help = 'Load data from Excel into CustomUser, Table, and Booking models'

    def handle(self, *args, **options):
        file_path = os.path.join(settings.BASE_DIR, 'tableapp', 'fixtures', 'data_ex.xlsx')

        # อ่านข้อมูลจาก Excel
        wb = load_workbook(file_path)

        # แผ่นงานสำหรับ CustomUser
        user_sheet = wb['CustomUser']
        for row in user_sheet.iter_rows(min_row=2, values_only=True):
            user_id, username, email, first_name, last_name, password = row

            if password and not isinstance(password, str):
               password = str(password)

            if password:
                user_instance, created = CustomUser.objects.get_or_create(
                    id=user_id,
                    defaults={
                        'username': username,
                        'email': email,
                        'first_name': first_name,
                        'last_name': last_name,
                    }
                )
                if created:
                    user_instance.set_password(password)
                    user_instance.save()
                    self.stdout.write(f"Created user: {username}")
                else:
                    self.stdout.write(f"User already exists: {username}")
            else:
                self.stdout.write(f"Skipping user {username} due to missing password")

        # แผ่นงานสำหรับ Table
        table_sheet = wb['Table']
        for row in table_sheet.iter_rows(min_row=2, values_only=True):
            table_id, table_name, table_status = row

            # สร้างหรืออัปเดตข้อมูล Table
            table_instance, created = Table.objects.get_or_create(
                id=table_id,
                defaults={
                    'table_name': table_name,
                    'table_status': table_status,
                }
            )
            if created:
                self.stdout.write(f"Created table: {table_name}")
            else:
                self.stdout.write(f"Table already exists: {table_name}")

        # แผ่นงานสำหรับ Booking
        booking_sheet = wb['Booking']
        for row in booking_sheet.iter_rows(min_row=2, values_only=True):
            booking_id, table_id, booking_date, booking_time, user_id = row

            # Debug: แสดงค่าของ row ก่อนประมวลผล
            self.stdout.write(f"Processing Booking row: {row}")

            # ตรวจสอบวันที่
            try:
                booking_date = datetime.strptime(str(booking_date), '%Y/%m/%d').date()
            except ValueError as e:
                self.stdout.write(f"Invalid booking date format for booking ID {booking_id}. Skipping... Error: {e}")
                continue

            # ตรวจสอบเวลา
            try:
                if isinstance(booking_time, datetime):
                    # ดึงแค่เวลาออกจาก datetime
                    booking_time = booking_time.time()
                elif isinstance(booking_time, time):
                    # ถ้าเป็นเวลาโดยตรง ใช้ได้เลย
                    pass
                else:
                    # หากเป็น string, แปลงเป็น time
                    booking_time = datetime.strptime(str(booking_time), '%H:%M:%S').time()
            except ValueError as e:
                self.stdout.write(f"Invalid booking time format for booking ID {booking_id}. Skipping... Error: {e}")
                continue

            # ค้นหา Table
            table = Table.objects.filter(id=table_id).first()
            if not table:
                self.stdout.write(f"Table ID {table_id} not found. Skipping booking.")
                continue

            # ค้นหา User
            user = None
            if user_id:
                user = CustomUser.objects.filter(id=user_id).first()
                if not user:
                    self.stdout.write(f"User ID {user_id} not found. Skipping booking ID {booking_id}.")

            # สร้างข้อมูลการจอง
            booking_instance, created = Booking.objects.get_or_create(
                id=booking_id,
                defaults={
                    'table': table,
                    'booking_date': booking_date,
                    'booking_time': booking_time,
                    'user': user,  # user อาจเป็น None
                }
            )
            if created:
                self.stdout.write(f"Created booking ID {booking_id} for table {table.table_name}")
            else:
                self.stdout.write(f"Booking ID {booking_id} already exists.")

        self.stdout.write(self.style.SUCCESS("Data loaded successfully!"))